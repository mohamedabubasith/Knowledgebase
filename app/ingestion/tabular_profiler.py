"""
Tabular document profiler.

Profiles CSV / XLSX schema using pure Python (no DuckDB).
Schema info (columns, types, samples, row count) is extracted locally,
then the file is uploaded to MindsDB for SQL query execution at search time.

Supported input types
---------------------
- text/csv  |  .csv
- text/tab-separated-values  |  .tsv
- application/vnd.openxmlformats-officedocument.spreadsheetml.sheet  |  .xlsx
- application/vnd.ms-excel  |  .xls  (first sheet only)
"""
import asyncio
import csv
import io
import os
import re
from typing import Optional

import structlog

log = structlog.get_logger(__name__)

# ---------------------------------------------------------------------------
# MIME / extension detection
# ---------------------------------------------------------------------------

TABULAR_MIMES = frozenset({
    "text/csv",
    "text/tab-separated-values",
    "application/csv",
    "application/vnd.ms-excel",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
})

TABULAR_EXTENSIONS = frozenset({".csv", ".tsv", ".xlsx", ".xls"})

_XLSX_MIMES = frozenset({
    "application/vnd.ms-excel",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
})


def is_tabular(filename: str, mime_type: str) -> bool:
    """Return True if this file should be handled by the tabular pipeline."""
    mime = mime_type.lower().split(";")[0].strip()
    ext = os.path.splitext(filename.lower())[1]
    return mime in TABULAR_MIMES or ext in TABULAR_EXTENSIONS


# ---------------------------------------------------------------------------
# XLSX → CSV conversion (pure Python, openpyxl)
# ---------------------------------------------------------------------------

def _xlsx_to_csv_bytes(data: bytes, sheet_name: Optional[str] = None) -> tuple[bytes, str]:
    """
    Convert an XLSX/XLS workbook to UTF-8 CSV bytes.
    Returns (csv_bytes, actual_sheet_name).
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    actual_sheet: str = ws.title  # type: ignore[assignment]

    buf = io.StringIO()
    writer = csv.writer(buf, quoting=csv.QUOTE_MINIMAL)
    for row in ws.iter_rows(values_only=True):  # type: ignore[union-attr]
        writer.writerow(["" if v is None else str(v) for v in row])
    wb.close()

    return buf.getvalue().encode("utf-8"), actual_sheet


# ---------------------------------------------------------------------------
# Type inference
# ---------------------------------------------------------------------------

_INT_PAT   = re.compile(r"^-?\d+$")
_FLOAT_PAT = re.compile(r"^-?\d*\.?\d+([eE][+-]?\d+)?$")


def _infer_type(samples: list[str]) -> str:
    non_empty = [s for s in samples if s.strip()]
    if not non_empty:
        return "VARCHAR"
    if all(_INT_PAT.match(s.strip()) for s in non_empty):
        return "BIGINT"
    if all(_FLOAT_PAT.match(s.strip()) for s in non_empty):
        return "DOUBLE"
    return "VARCHAR"


# ---------------------------------------------------------------------------
# Pure-Python CSV profiling
# ---------------------------------------------------------------------------

def _profile_sync(filename: str, data: bytes, mime_type: str) -> tuple[dict, bytes]:
    """
    Profile a tabular file using pure Python.

    Returns (schema_dict, csv_bytes) where csv_bytes is ready for MindsDB upload.
    XLSX files are converted to CSV bytes first.

    Schema dict:
        {
            "columns": [{"name": str, "type": str, "sample": [str, ...]}, ...],
            "row_count": int,
            "sheet_name": str | None,
        }
    """
    mime = mime_type.lower().split(";")[0].strip()
    ext  = os.path.splitext(filename.lower())[1]

    sheet_name: Optional[str] = None
    csv_bytes = data

    if mime in _XLSX_MIMES or ext in (".xlsx", ".xls"):
        csv_bytes, sheet_name = _xlsx_to_csv_bytes(data)

    csv_text = csv_bytes.decode("utf-8", errors="replace")
    reader   = csv.reader(io.StringIO(csv_text))

    try:
        headers = next(reader)
    except StopIteration:
        return {"columns": [], "row_count": 0, "sheet_name": sheet_name}, csv_bytes

    sample_rows: list[list[str]] = []
    row_count = 0
    for row in reader:
        row_count += 1
        if row_count <= 5:
            sample_rows.append(row)

    columns = []
    for i, h in enumerate(headers):
        samples = [
            str(row[i]) for row in sample_rows
            if i < len(row) and row[i] is not None and str(row[i]).strip()
        ][:5]
        columns.append({"name": h, "type": _infer_type(samples), "sample": samples})

    return {
        "columns":    columns,
        "row_count":  row_count,
        "sheet_name": sheet_name,
    }, csv_bytes


# ---------------------------------------------------------------------------
# Async wrapper
# ---------------------------------------------------------------------------

async def profile_tabular(filename: str, data: bytes, mime_type: str) -> tuple[dict, bytes]:
    """
    Async wrapper around `_profile_sync`.

    Returns (schema_dict, csv_bytes).
    csv_bytes is suitable for direct upload to MindsDB.
    """
    loop = asyncio.get_event_loop()
    schema, csv_bytes = await loop.run_in_executor(None, _profile_sync, filename, data, mime_type)
    log.info(
        "tabular_profiled",
        filename=filename,
        columns=len(schema["columns"]),
        row_count=schema["row_count"],
        sheet=schema.get("sheet_name"),
    )
    return schema, csv_bytes


# ---------------------------------------------------------------------------
# Summary chunk builder
# ---------------------------------------------------------------------------

def build_summary_chunk(filename: str, schema: dict) -> str:
    """
    Build a human-readable summary of the tabular file for embedding + FTS.
    """
    cols      = schema.get("columns", [])
    row_count = schema.get("row_count", 0)
    sheet     = schema.get("sheet_name")

    col_desc = ", ".join(
        f"{c['name']} ({c['type']})" for c in cols[:30]
    )

    sample_lines = []
    for c in cols[:5]:
        if c.get("sample"):
            sample_lines.append(
                f"  • {c['name']}: {', '.join(c['sample'][:3])}"
            )
    sample_text = ("\nSample values:\n" + "\n".join(sample_lines)) if sample_lines else ""

    sheet_info = f" Sheet: {sheet}." if sheet else ""
    return (
        f"Tabular data file: {filename}.{sheet_info} "
        f"Contains {row_count:,} rows with {len(cols)} columns.\n"
        f"Columns: {col_desc}.{sample_text}\n"
        f"This file supports analytics queries such as aggregations, filtering, "
        f"grouping, trends, comparisons, totals, averages, and counts."
    )
